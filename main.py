import os
import glob
import openpyxl

archivo_principal = "G:\\ReadWriteFile\\ReadWriteFile\\template\\FORMATO-COVID.xlsx"
wb_principal = openpyxl.load_workbook(archivo_principal)

nueva_linea_triage = ["F", "G", "H", "I", "J", "K", "L", "M"]  # Celdas en la hoja "TRIAJE"
nueva_linea_odontologia = ["B", "C", "D", "E", "F", "G", "H", "I"]  # Celdas en la hoja "REG.ODONTOLOGIA"

hoja_triage = wb_principal["TRIAGE"]
hoja_odontologia = wb_principal["REG.ODONTOLOGIA"]

# Obtener el valor de las celdas de la nueva línea en TRIAJE
valores_triage = [hoja_triage[cell][0].value for cell in nueva_linea_triage]

# Verificar si la nueva línea ya existe en REG.ODONTOLOGIA
existe_linea = False
for row in hoja_odontologia.iter_rows(min_row=2, values_only=True):
    if row[:len(valores_triage)] == valores_triage:
        existe_linea = True
        break

if existe_linea:
    print("La nueva línea ya existe en la hoja REG.ODONTOLOGIA.")
else:
    # Insertar la nueva línea en REG.ODONTOLOGIA
    row_index = hoja_odontologia.max_row + 1
    for i, valor in enumerate(valores_triage):
        cell = nueva_linea_odontologia[i] + str(row_index)
        hoja_odontologia[cell] = valor

    print("La nueva línea se ha agregado correctamente a la hoja REG.ODONTOLOGIA.")

    # Mostrar los datos agregados en REG.ODONTOLOGIA
    datos_agregados = [hoja_odontologia[cell][0].value for cell in nueva_linea_odontologia]
    print("Datos agregados en REG.ODONTOLOGIA:", datos_agregados)

# Guardar los cambios en el archivo
wb_principal.save(archivo_principal)

# Borrar archivos temporales generados por Openpyxl
temp_files = glob.glob(os.path.join(os.getcwd(), "~$*.xlsx"))
for file in temp_files:
    os.remove(file)
